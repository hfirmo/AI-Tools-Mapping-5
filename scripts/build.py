#!/usr/bin/env python3
"""Rebuild output/mapping.md and output/mapping.xlsx from the two source layers.

Editorial layer:  data/editorial.md   (human judgments - edit freely)
Volatile layer:   data/volatile.json  (models/prices - updated via refresh.py + review)
Never edit files in output/ by hand; they are overwritten on every build.
"""
import json, re, datetime, pathlib

ROOT = pathlib.Path(__file__).resolve().parents[1]
ED = (ROOT / "data" / "editorial.md").read_text(encoding="utf-8")
VOL = json.loads((ROOT / "data" / "volatile.json").read_text(encoding="utf-8"))
OUT = ROOT / "output"
OUT.mkdir(exist_ok=True)

def key_of(tool_cell: str) -> str:
    return re.sub(r"\s*\*\(.*?\)\*", "", tool_cell).strip()

stamp = datetime.date.today().isoformat()
lines_out, in_t1 = [], False
for ln in ED.splitlines():
    s = ln.strip()
    if s.startswith("| Category | Tool | Owner |"):
        in_t1 = True
        lines_out.append(
            "| Category | Tool | Owner | Latest model | Access pricing | "
            "API price per 1M tokens in/out | Contribution to QDA | Strengths | Limitations | Order |")
        continue
    if in_t1 and s.startswith("|---"):
        lines_out.append("|" + "---|" * 10)
        continue
    if in_t1 and s.startswith("|"):
        p = ln.split("|")
        v = VOL.get(key_of(p[2]), {})
        lm = v.get("latest_model", "—")
        ap = v.get("access_pricing", "—")
        tp = v.get("token_price", "—")
        if v and v.get("status") != "reviewed":
            tp += "  ⚠ unreviewed auto-update"
        p.insert(4, " " + tp + " ")
        p.insert(4, " " + ap + " ")
        p.insert(4, " " + lm + " ")
        lines_out.append("|".join(p))
        continue
    if in_t1 and not s.startswith("|"):
        in_t1 = False
    lines_out.append(ln)

retrieved = sorted({v.get("retrieved", "?") for v in VOL.values()})
banner = (f"\n> **Generated document — do not edit.** Built {stamp} from data/editorial.md + "
          f"data/volatile.json (volatile data retrieved: {', '.join(retrieved)}). "
          f"Volatile cells carry per-entry sources in data/volatile.json.\n")
text = "\n".join(lines_out)
text = text.replace("\nScope of analysis follows", banner + "\nScope of analysis follows", 1)
(OUT / "mapping.md").write_text(text, encoding="utf-8")

# ---------- Excel ----------
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def clean(c):
    return re.sub(r"\s+", " ", c.replace("**", "").replace("*", "")).strip()

tables, cur = [], []
for ln in text.splitlines():
    if ln.strip().startswith("|"):
        cur.append(ln)
    elif cur:
        tables.append(cur); cur = []
if cur: tables.append(cur)

def parse(tbl):
    rows = [[clean(c) for c in ln.strip().strip("|").split("|")] for ln in tbl]
    return rows[0], rows[2:]

t1h, t1b = parse(tables[0]); t2h, t2b = parse(tables[1]); t3h, t3b = parse(tables[2])
last = ""
for r in t1b:
    if r[0]: last = r[0]
    else: r[0] = last

FONT = "Arial"
hfill = PatternFill("solid", fgColor="1F3864")
bfill = PatternFill("solid", fgColor="EDF2F9")
thin = Side(style="thin", color="C9C9C9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wb = Workbook()

def sheet(ws, header, body, widths, band=False):
    ws.append(header)
    for j in range(1, len(header) + 1):
        c = ws.cell(1, j); c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        c.fill = hfill; c.alignment = Alignment(vertical="center", wrap_text=True); c.border = border
    prev, on = None, False
    for r in body:
        ws.append(r); i = ws.max_row
        if band and r[0] != prev: on, prev = not on, r[0]
        for j in range(1, len(header) + 1):
            c = ws.cell(i, j); c.font = Font(name=FONT, size=10)
            c.alignment = Alignment(vertical="top", wrap_text=True); c.border = border
            if band and on: c.fill = bfill
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"

ws1 = wb.active; ws1.title = "Tool landscape"
sheet(ws1, t1h, t1b, [22, 24, 24, 30, 26, 30, 38, 42, 46, 12], band=True)
sheet(wb.create_sheet("Build-your-own ladder"), t2h, t2b, [8, 48, 48, 48])
sheet(wb.create_sheet("Research communities"), t3h, t3b, [42, 46, 50])
ws0 = wb.create_sheet("Legend & notes", 0)
ws0["A1"] = "GenAI Tools for QDA — generated " + stamp
ws0["A1"].font = Font(name=FONT, bold=True, size=13)
ws0["A2"] = "Built from editorial.md + volatile.json; per-cell sources and retrieval dates in volatile.json."
ws0["A2"].font = Font(name=FONT, italic=True, size=9, color="666666")
ws0.column_dimensions["A"].width = 110

# ---------- optional appendix: comprehensive model index ----------
mi_path = ROOT / "data" / "model_index.json"
if mi_path.exists():
    mi = json.loads(mi_path.read_text(encoding="utf-8") or "{}")
    if not mi.get("models"):
        print("Model index file present but empty — sheet skipped.")
        mi = None
if mi_path.exists() and mi:
    ws4 = wb.create_sheet("Model index (auto)")
    hdr = ["Provider", "Model key", "Input $/MTok", "Output $/MTok", "Context window"]
    body = [[m["provider"], m["model"], m["input_per_mtok"], m["output_per_mtok"], m["context_window"]]
            for m in mi.get("models", [])]
    sheet(ws4, hdr, body, [22, 46, 16, 16, 18])
    ws4.insert_rows(1)
    ws4["A1"] = (f'Auto-generated appendix — {mi.get("count", 0)} text models, retrieved '
                 f'{mi.get("retrieved", "?")}. Source: {mi.get("source", "?")}. '
                 f'List prices only, no analytic judgment; quality rankings: artificialanalysis.ai/models')
    ws4["A1"].font = Font(name=FONT, italic=True, size=9, color="666666")
    ws4.freeze_panes = "A3"
    print(f"Appendix sheet: {mi.get('count', 0)} models")


# ---------- optional appendix: Artificial Analysis benchmarks (needs AA_API_KEY) ----------
aa_path = ROOT / "data" / "aa_index.json"
if aa_path.exists():
    aa = json.loads(aa_path.read_text(encoding="utf-8") or "{}")
    if not aa.get("models"):
        print("AA file present but empty — sheet skipped.")
        aa = None
if aa_path.exists() and aa:
    ws5 = wb.create_sheet("AA benchmarks (auto)")
    hdr = ["Model", "Creator", "Intelligence Index", "Output speed (t/s)", "Latency (s)",
           "Input $/MTok", "Output $/MTok", "Context window"]
    body = [[m.get("model"), m.get("creator"), m.get("intelligence_index"),
             m.get("output_speed_tps"), m.get("latency_s"), m.get("price_input_per_mtok"),
             m.get("price_output_per_mtok"), m.get("context_window")]
            for m in aa.get("models", [])]
    sheet(ws5, hdr, body, [38, 22, 16, 16, 12, 14, 14, 16])
    ws5.insert_rows(1)
    ws5["A1"] = (f'Auto-generated {aa.get("retrieved","?")} — {aa.get("count",0)} models. '
                 f'{aa.get("source","")}. Benchmark data © Artificial Analysis; cite artificialanalysis.ai.')
    ws5["A1"].font = Font(name=FONT, italic=True, size=9, color="666666")
    ws5.freeze_panes = "A3"
    print(f"AA sheet: {aa.get('count', 0)} models")

wb.save(OUT / "mapping.xlsx")
print("Built", OUT / "mapping.md", "and", OUT / "mapping.xlsx")
